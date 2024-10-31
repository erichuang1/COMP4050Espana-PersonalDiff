# JSON -> MD formatting
#
# by Harris C. McRae, Eric
#
# Formats dicts for rubrics & viva questions into a MD format string.

from jinja2 import Environment, FileSystemLoader
import openpyxl
import io
import pandas as pd
import difflib

def rubric_make_md(data:dict)->str:
    """Formats a MD string from a rubric dict type."""
    output = ''

    if data.get('rubric_title') is not None:
        output += f"# {data['rubric_title']}\n"

    if data.get("grade_descriptors") is None:
        return output

    gds = data["grade_descriptors"]

    keys = [ "fail", "pass_", "credit", "distinction", "high_distinction" ]
    output += f"|   | Fail | Pass | Credit | Distinction | High Distinction |\n"
    output += f"| :- | :-: | :-: | :-: | :-: | :-: |\n"

    output += f"| **Grade%** "
    for key in keys:
        grade = gds[key]
        output += f"| {grade['mark_min']}-{grade['mark_max']}%"

    output += "|\n"

    sentinel = True
    index = 0
    while sentinel:
        added = False
        for key in keys:
            grade = gds[key]
            criteria = grade["criterion"]
            if not index < len(criteria):
                sentinel = False
                break
            specific_criteria = criteria[index]

            if not added:
                output += f"| **{specific_criteria['criteria_name']}** "
                added = True

            output += f"| {specific_criteria['criteria_description']} "

        if not sentinel:
            break
        output += "| \n"
        index += 1

    return output

def rubric_make_xls_data(data:dict)->(bool,io.BytesIO):
    """Formats an XLS file from a given rubric."""
    
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    
    if data.get("grade_descriptors") is None:
        return False, None

    col_headers = ['a', 'b', 'c', 'd', 'e', 'f']
    keys = [ "fail", "pass_", "credit", "distinction", "high_distinction" ]
    gds = data["grade_descriptors"]
    
    # add data to workbook
    xls_data = [
            [ '', 'Fail', 'Pass', 'Credit', 'Distinction', 'High Distinction' ]
        ]

    tmp_list:str = ['']
    
    for key in keys:
        grade = gds[key]
        tmp_list.append(f"{grade['mark_min']}-{grade['mark_max']}%")

    xls_data.append(tmp_list)

    sentinel = True
    index = 0
    while sentinel:
        tmp_list = []
        added = False
        for key in keys:
            grade = gds[key]
            criteria = grade["criterion"]
            if not index < len(criteria):
                sentinel = False
                break
            specific_criteria = criteria[index]

            if not added:
                tmp_list.append(f"{specific_criteria['criteria_name']}")
                added = True

            tmp_list.append(f"{specific_criteria['criteria_description']}")

        if not sentinel:
            break
        
        xls_data.append(tmp_list)
        index += 1

    for row in xls_data:
        sheet.append(row)

    width = len(col_headers)+1
    height = len(xls_data)+1

    # control width & height of all columns
    for col in col_headers:
        if col == 'a':
            sheet.column_dimensions[col].width = 30
        else:
            sheet.column_dimensions[col].width = 50
    for i in range(3, height):
        sheet.row_dimensions[i].height = 80

    # style top row
    for k in range(1, width):
        sheet.cell(row=1, column=k).font = openpyxl.styles.Font(bold=True, size=12)
        sheet.cell(row=1, column=k).alignment = openpyxl.styles.Alignment(vertical='top',horizontal='center')

    # style second row
    for k in range(1, width):
        if k == 1:
            sheet.cell(row=2, column=k).font = openpyxl.styles.Font(bold=True, size=12)
        else:
            sheet.cell(row=2, column=k).font = openpyxl.styles.Font(bold=False, size=12)
        sheet.cell(row=2, column=k).alignment = openpyxl.styles.Alignment(vertical='top',horizontal='center')

    # style left column headers
    for i in range(3, height):
        sheet.cell(row=i, column=1).font = openpyxl.styles.Font(bold=True, size=12)
        sheet.cell(row=i, column=1).alignment = openpyxl.styles.Alignment(vertical='center',horizontal='center')

    # style remaining data cells
    for i in range(3, height):
        for k in range(2, width):
            sheet.cell(row=i, column=k).font = openpyxl.styles.Font(bold=False, size=12)
            sheet.cell(row=i, column=k).alignment = openpyxl.styles.Alignment(vertical='center',horizontal='center',wrap_text=True,shrink_to_fit=True)

    # set borders
    for i in range(1, height):
        for k in range(1, width):
            sheet.cell(row=i,column=k).border = openpyxl.styles.Border(left=openpyxl.styles.Side('thin'),right=openpyxl.styles.Side('thin'), \
                                                                       top=openpyxl.styles.Side('thin'),bottom=openpyxl.styles.Side('thin'))
    
    out_bytes = io.BytesIO()
    workbook.save(out_bytes)

    return True,out_bytes

_AI_QN_CATEGORIES = ['analysis_evaluation', 'application_problem_solving', 'factual_recall', 'open_ended', 'conceptual_understanding']
_AI_QN_CATG_PROPER_NAME = {
        'analysis_evaluation' : 'Analysis & Evaluation',
        'application_problem_solving' : 'Application & Problem Solving',
        'factual_recall' : 'Factual Recall',
        'open_ended' : 'Open-Ended',
        'conceptual_understanding' : 'Conceptual understanding'
    }

def viva_make_md(data:dict)->str:
    """Formats a MD string from a viva question dict."""

    output = ''
    
    output += f"# {str(data['project_title'])}\n"
    output += f"## Submission ID #{str(data['submission_id'])} | {str(data['unit_code'])}\n"

    output += "\n"
    output += f"## Static Questions\n"
    # Static Questions append

    idx = 1
    for qn in data['static_questions']:
        output += f"{idx}. {qn}\n"
        idx += 1
    
    output += "\n"
    output += f"## Random Questions\n"
    # Random Questions append

    rqn = data['random_questions']

    idx = 1
    for question in rqn:
        output += f"{idx}. {question['question']}\n"
        idx += 1

    output += "\n"
    output += f"## AI Questions\n"
    # AI questions append
    
    temp = data['ai_questions']
    for key in _AI_QN_CATEGORIES:
        ai_qns = temp[key]

        idx = 1
        while True:
            question = ai_qns.get(f"question_{idx}")
            if question is None:
                break

            if idx == 1:
                output += f"### {_AI_QN_CATG_PROPER_NAME[key]} questions:\n"

            output += f"{idx}. {question}\n"
            idx += 1

    return output

_sample_rubric = {
    "rubric_title": "Operating Systems Assessment Marking Rubric",
    "grade_descriptors": {
        "fail": {
            "mark_min": 0,
            "mark_max": 49,
            "criterion": [
                {
                    "criteria_name": "Process Management",
                    "criteria_description": "Inadequate understanding of process lifecycle and synchronization. No application of scheduling algorithms."
                },
                {
                    "criteria_name": "Memory Management",
                    "criteria_description": "Little to no comprehension of memory allocation strategies. Cannot apply virtual memory concepts."
                },
                {
                    "criteria_name": "File Systems",
                    "criteria_description": "Insufficient knowledge of file system design and optimization techniques."
                }
            ]
        },
        "pass_": {
            "mark_min": 50,
            "mark_max": 64,
            "criterion": [
                {
                    "criteria_name": "Process Management",
                    "criteria_description": "Basic understanding of process lifecycle. Some discussion of scheduling but lacks depth."
                },
                {
                    "criteria_name": "Memory Management",
                    "criteria_description": "Basic knowledge of memory allocation strategies, with limited application of virtual memory concepts."
                },
                {
                    "criteria_name": "File Systems",
                    "criteria_description": "Basic understanding of file system architecture and file allocation methods."
                }
            ]
        },
        "credit": {
            "mark_min": 65,
            "mark_max": 74,
            "criterion": [
                {
                    "criteria_name": "Process Management",
                    "criteria_description": "Good understanding of process lifecycle and scheduling with examples. Some skills in solving synchronization issues."
                },
                {
                    "criteria_name": "Memory Management",
                    "criteria_description": "Competently apply virtual memory concepts and describe paging and segmentation."
                },
                {
                    "criteria_name": "File Systems",
                    "criteria_description": "Ability to evaluate file system design and apply knowledge of disk scheduling."
                }
            ]
        },
        "distinction": {
            "mark_min": 75,
            "mark_max": 84,
            "criterion": [
                {
                    "criteria_name": "Process Management",
                    "criteria_description": "Strong understanding of scheduling algorithms and process synchronization. Effectively resolves synchronization issues."
                },
                {
                    "criteria_name": "Memory Management",
                    "criteria_description": "Demonstrates detailed knowledge of memory allocation strategies and consistently applies virtual memory concepts."
                },
                {
                    "criteria_name": "File Systems",
                    "criteria_description": "Analytical evaluation of file system architecture; optimizes file systems with sound methodologies."
                }
            ]
        },
        "high_distinction": {
            "mark_min": 85,
            "mark_max": 100,
            "criterion": [
                {
                    "criteria_name": "Process Management",
                    "criteria_description": "Exceptional comprehension of process management concepts. Innovatively solves complex synchronization problems."
                },
                {
                    "criteria_name": "Memory Management",
                    "criteria_description": "Mastery of memory management principles; expertly applies advanced paging and segmentation strategies."
                },
                {
                    "criteria_name": "File Systems",
                    "criteria_description": "Creates innovative solutions in file system design and optimization, backed by thorough analysis."
                }
            ]
        }
    }
}
_sample_viva = {
        "ai_questions": {
            "analysis_and_evaluation": {
                "question_1": "Evaluate the importance of documentation in a project like yours that involves complex programming tasks. How will you ensure your documentation remains consistent and helpful?",
                "question_2": "Analyze the decision to use C++ for developing your program. What are the advantages and disadvantages of using C++ for low-level graphics programming compared to other languages?",
                "question_3": "In your opinion, how does the complexity of the Vulkan API compare to other graphics libraries you are aware of? What challenges do you anticipate facing with Vulkan?"
            },
            "application_and_problem_solving": {
                "question_1": "Given your understanding of graphics programming, how would you approach implementing a new shader effect that is not covered in the existing resources you plan to use?"
            },
            "factual_recall": {
                "question_1": "What is the primary purpose of learning the Vulkan graphics library and GLSL in the context of your project?",
                "question_2": "Can you explain the basic differences between Vulkan and OpenGL?",
                "question_3": "What are the main milestones outlined in your project, and what does each milestone aim to achieve?"
            },
            "open_ended": {
                "question_1": "Reflecting on your project, how do you envision the skills you develop through this project impacting your future career in game development or related fields?"
            }
        },
        "project_title": "Project Himanshi",
        "random_questions": [
            {
                "question": "q1"
            },
            {
                "question": "q2"
            },
            {
                "question": "q3"
            }
        ],
        "static_questions": [
            "Who are you",
            "What is your favourite brownie flavour"
        ],
        "submission_id": 10,
        "unit_code": "CS101"
    }

def correct_ai_output(data:dict)->dict:
    """Corrects any misspelled AI categories to match field _AI_QN_CATEGORIES"""
    
    output:dict = { }
    
    for tgt_key in _AI_QN_CATEGORIES:
        h_match = -1
        h_key = ''
        # For each key in the categories list, find the key in data that matches. Delete the key from data once done.
        for src_key in data.keys():
            score = get_similarity(src_key, tgt_key)
            if score > h_match:
                h_match = score
                h_key = src_key
        
        if h_match != -1:
            print(f'Correcting {h_key} to {tgt_key}\ndata:{data[h_key]}')
            output[tgt_key] = data[h_key]
            data.pop(h_key, None)
        
    return output

def get_similarity(input_string, reference_string):
    #The ndiff method returns a list of strings representing the differences between the two input strings.
    diff = difflib.ndiff(input_string, reference_string)
    diff_count = 0
    for line in diff:
      # a "-", indicating that it is a deleted character from the input string.
        if line.startswith("-"):
            diff_count += 1
    # calculates the similarity by subtracting the ratio of the number of deleted characters to the length of the input string from 1
    return 1 - (diff_count / len(input_string))